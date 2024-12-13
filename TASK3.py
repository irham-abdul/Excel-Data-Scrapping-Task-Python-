import os
import pandas as pd                             # pip install pandas
from openpyxl.styles import Font
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook    # pip install openpyxl
import xlwings as xw                            # pip install xlwings


def convert_xlsb2xlsx(filepath):
    xlsx_path = filepath.replace('.xlsb', '.xlsx')
    with xw.App(visible=False) as app:
        wb = app.books.open(filepath)
        wb.save(xlsx_path)
        wb.close()
    return xlsx_path

def extract_and_append_rows(source_file, target_file, source_sheet_name, target_sheet_name, source_row_start_index, target_column):
        # Check the file extension to decide method for reading the file
        file_extension = os.path.splitext(source_file)[1].lower()

        # If file is .xlsb, convert it to .xlsx before extracting data
        if file_extension == '.xlsb':
            xlsx_file = convert_xlsb2xlsx(source_file)  # Convert .xlsb to .xlsx
            source_file = xlsx_file  # Update the source file to the converted .xlsx file

        # Reading the .xlsx file (both new and converted)
        df_source = pd.read_excel(source_file, sheet_name=source_sheet_name, index_col=None)
        source_wb = load_workbook(source_file)

        # Get the value from cell B2 in the first sheet of .xlsx
        value_b2 = source_wb.worksheets[0]['B2'].value  # Accessing the first sheet (index 0)

        if value_b2 is None:
            value_b2 = "No value in B2"
        
        # Get name of the report from the source file without the extension
        report_name = os.path.basename(source_file).split('.')[0]  

        # Get the absolute path of the source file
        source_file_path = os.path.abspath(source_file)
        
        # Load the target workbook and worksheet
        target_wb = load_workbook(target_file)
        target_ws = target_wb[target_sheet_name]

        # Identify the last used row in the target sheet, start appending from row 2 onward
        last_row = target_ws.max_row  # Get the last used row in the sheet (including the header)

        # If the last row is 1 (only the header), then start from row 2
        if last_row == 1:
            last_row = 1
        else:
            last_row = target_ws.max_row + 1  # Set last_row to the next row after the header

        # Insert the report name into column B (starting from row 2)
        for row in range(last_row, last_row + len(df_source)):  # Adjusted to start at last_row
            target_ws.cell(row=row, column=2, value=report_name)  

        # Insert the source file path into column C
        for row in range(last_row, last_row + len(df_source)):
            target_ws.cell(row=row, column=3, value=source_file_path)  

        # Insert the value from B2 of the first sheet into column D
        for row in range(last_row, last_row + len(df_source)):
            target_ws.cell(row=row, column=4, value=value_b2)  

        # Calculate yesterday's date in the required format (dd/mm/yyyy)
        yesterday_date = (datetime.now() - timedelta(1)).strftime('%d/%m/%Y')

        # Insert "zsystem" header and yesterday's date in the first row only (column L)
        if last_row == 1:  # Only if this is the first time appending (header is empty)
            bold_font = Font(bold=True)
            target_ws.cell(row=1, column=12, value="zsystem").font = bold_font  # Add header in L1 (bold font)
            target_ws.cell(row=2, column=12, value=yesterday_date)  # Add date in L2
        
        # Loop through each row in the source, starting from the specified start index
        for row_offset in range(len(df_source)):
            row_data = df_source.iloc[row_offset, 3:8]  # Extract relevant data from columns 4 to 8
            
            target_row = last_row + row_offset  # Ensure data is appended after the last row
            # Loop through columns to populate data
            for col_offset, value in enumerate(row_data):
                target_ws.cell(row=target_row, column=7 + col_offset, value=value)
                
            # Insert CONCATENATE formula into Column F for current row, dynamically using value from column 5 (E)
            target_ws.cell(
                row=target_row, 
                column=6,  # Column F is column 6 (1-based index)
                value=f'=CONCATENATE("{target_ws.cell(row=target_row, column=5).value} as at ","-",TEXT(L2,"dd-mmm-yyyy"))'
            )

        # Save updated target workbook (overwriting any previous data if necessary)
        target_wb.save(target_file)
        return True  # Return True if data extraction and appending are successful 

# Function to read file paths from the text file
def get_source_files(file_path):
    try:
        with open(file_path, 'r') as f:
            # Read all lines from the text file and strip any extra spaces or newlines
            file_paths = [line.strip() for line in f.readlines()]
        return file_paths
    except FileNotFoundError:
        print(f"Text file not found: {file_path}")
        return []

# File paths, sheet names, and row/column details for target
source_files = get_source_files(r"C:\Users\mirham\Downloads\INTERN FILE\TASK 3\PART 2\Report_Template_Paths.txt")  # Read the source files from the text file
target_file = r"C:\Users\mirham\Downloads\INTERN FILE\TASK 3\PART 2\Final Extraction.xlsx"
source_sheet_name = "Rpt-Maintain"  
target_sheet_name = "Sheet1"         
source_row_start_index = 0           # Start at the first row in the source sheet, Base 0
target_column = 4                    # Starting column to insert into in the target sheet, 1-based index

# Initialize a list to track successful extractions
successful_extractions = []

# Check if target file exists. If it exists, we will clear data from row 2 onward before appending new data
if os.path.exists(target_file):
    # Load the target workbook to clear its content
    target_wb = load_workbook(target_file)
    target_ws = target_wb[target_sheet_name]

    # Clear the content of the target sheet starting from row 2 onward (keeping header intact)
    for row in target_ws.iter_rows(min_row=2, max_row=target_ws.max_row, min_col=1, max_col=target_ws.max_column):
        for cell in row:
            cell.value = None

    target_wb.save(target_file)

# Process each source file in the list
for idx, source_file in enumerate(source_files):
    if extract_and_append_rows(source_file, target_file, source_sheet_name, target_sheet_name, source_row_start_index, target_column):
        successful_extractions.append(f"Finished extracting data from report {idx + 1}")
    else:
        successful_extractions.append(f"Failed to extract data from report {idx + 1}")

# Print summary after all reports have been processed
print("\n".join(successful_extractions))

# After processing, output the final extraction data to text and csv
df = pd.read_excel(target_file, sheet_name=target_sheet_name)  # target file to write extracted data
df.to_csv(r'C:\Users\mirham\Downloads\INTERN FILE\TASK 3\PART 2\Final_Extraction.txt', sep='\t', index=False)  # Portion to write from excel file to text.
print("Successfully output excel data to Final_Extraction.txt") 
df.to_csv(r'C:\Users\mirham\Downloads\INTERN FILE\TASK 3\PART 2\Final_Extraction.csv', sep='\t', index=False)  # Portion to write from excel file to csv.
print("Successfully output excel data to Final_Extraction.csv")
