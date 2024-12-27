import os
import openpyxl
import xlwings as xw
from datetime import datetime, timedelta
import psutil  # Module for managing system processes (closing any excel file before start of operation)

# Paths
txt_file_path = r"Report_Template_Paths.txt"
destination_file_path = r"Final Extraction.xlsx"
converted_dir = r"Converted"

# Ensure the Converted folder exists
os.makedirs(converted_dir, exist_ok=True)

# Function to terminate any open Excel processes
def close_open_excel_files():
    """
    Closes any open Excel files before starting the process.
    """
    print("Closing any open Excel files...")
    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] and process.info['name'].lower() == 'excel.exe':
            try:
                process.terminate()
                process.wait(timeout=5)
                print(f"Terminated: {process.info['name']}")
            except Exception as e:
                print(f"Failed to terminate {process.info['name']}: {e}")

# Function to convert .xlsb to .xlsx
def convert_xlsb_to_xlsx(xlsb_file_path):
    """
    Converts an .xlsb file to .xlsx format using xlwings and saves it in the Converted folder.
    """
    xlsx_file_name = os.path.basename(xlsb_file_path).replace('.xlsb', '.xlsx')
    xlsx_file_path = os.path.join(converted_dir, xlsx_file_name)

    try:
        with xw.App(visible=False) as app:
            wb = app.books.open(xlsb_file_path)
            app.enable_events = False
            wb.save(xlsx_file_path)
            wb.close()
        print(f"Converted: {xlsb_file_path} -> {xlsx_file_path}")
    except Exception as e:
        print(f"Error converting {xlsb_file_path} to .xlsx: {e}")
        raise

    if not os.path.exists(xlsx_file_path):
        raise FileNotFoundError(f"Converted file not found: {xlsx_file_path}")

    return xlsx_file_path

# Function to process files
def process_files(file_paths):
    """
    Processes a list of file paths and appends extracted data to the destination workbook.
    """
    if not os.path.exists(destination_file_path):
        print(f"Creating new destination file at {destination_file_path}")
        destination_wb = openpyxl.Workbook()
        destination_ws = destination_wb.active
    else:
        print(f"Opening existing destination file at {destination_file_path}")
        destination_wb = openpyxl.load_workbook(destination_file_path)
        destination_ws = destination_wb.active

    destination_ws.delete_rows(1, destination_ws.max_row)
    header = [
        "Report Template", "Template Path", "Report Path", "Report Name", 
        "Origin Value", "Filter", "Text", "Report Format", "Frequency", 
        "Term", "zsystem"
    ]
    destination_ws.append(header)
    for cell in destination_ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for file_path in file_paths:
        file_path = file_path.strip()
        if os.path.isdir(file_path):
            for root, _, files in os.walk(file_path):
                for file in files:
                    if file.startswith('~$'):  # Skip temporary files
                        print(f"Skipping temporary file: {file}")
                        continue
                    if file.endswith(('.xlsx', '.xlsb')):
                        full_file_path = os.path.join(root, file)
                        print(f"Processing file: {full_file_path}")
                        try:
                            process_excel_file(full_file_path, destination_ws)
                        except Exception as e:
                            print(f"Error processing file {full_file_path}: {e}")
        elif os.path.exists(file_path):
            if file_path.startswith('~$'):
                print(f"Skipping temporary file: {file_path}")
                continue
            if file_path.endswith(('.xlsx', '.xlsb')):
                try:
                    print(f"Processing file: {file_path}")
                    process_excel_file(file_path, destination_ws)
                except Exception as e:
                    print(f"Error processing file {file_path}: {e}")
            else:
                print(f"Skipping unsupported file: {file_path}")
        else:
            print(f"File or folder does not exist: {file_path}")

    destination_wb.save(destination_file_path)
    print("Data extraction complete.")

# Function to process an individual Excel file
def process_excel_file(file_path, destination_ws):
    """
    Processes an individual Excel file and appends its data to the destination sheet.
    """
    if file_path.endswith('.xlsb'):
        print(f"Converting .xlsb file: {file_path}")
        try:
            file_path = convert_xlsb_to_xlsx(file_path)
        except Exception as e:
            print(f"Failed to convert .xlsb file: {file_path}. Error: {e}")
            return

    if not file_path.endswith('.xlsx'):
        print(f"Skipping unsupported file: {file_path}")
        return

    try:
        content_wb = openpyxl.load_workbook(file_path, data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"Invalid Excel file: {file_path}. Skipping. Error: {e}")
        return
    except Exception as e:
        print(f"Unexpected error loading workbook {file_path}: {e}")
        return

    sheet_1, sheet_2 = content_wb.worksheets[:2]
    value_from_b2 = sheet_1['B2'].value
    rows_count_sheet_1, rows_count_sheet_2 = sheet_1.max_row, sheet_2.max_row
    previous_date = (datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")

    for i in range(2, max(rows_count_sheet_1, rows_count_sheet_2) + 1):
        row_data = [os.path.splitext(os.path.basename(file_path))[0], file_path, value_from_b2]
        row_data.append(sheet_2[f"C{i}"].value if i <= rows_count_sheet_2 else None)

        source_content = sheet_2[f"C{i}"].value if i <= rows_count_sheet_2 else None
        if source_content and "-" in source_content:
            string_part = source_content.split("-")[0].strip()
            row_data.append(
                f'=CONCATENATE("{string_part}", " at ", "-", TEXT(K2, "dd-mmm-yyyy"))'
            )
        else:
            row_data.append(
                '=CONCATENATE("Invalid Content in C", " at ", "-", TEXT(K2, "dd-mmm-yyyy"))'
            )

        for col in ['D', 'E', 'F', 'G', 'H']:
            row_data.append(sheet_2[f"{col}{i}"].value if i <= rows_count_sheet_2 else None)
        destination_ws.append(row_data)

    destination_ws['K2'] = previous_date

# Close any open Excel files before starting the process
close_open_excel_files()

# Read file paths from the text file
with open(txt_file_path, "r", encoding="utf-8") as file:
    file_paths = file.readlines()

# Process the files (both individual files and files inside folders)
process_files(file_paths)
